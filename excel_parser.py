import openpyxl
import json
import os
import re

EXCEL_PATH = r"C:\Users\llboz\OneDrive\Desktop\Agustos_2026_Uretim_Raporu.xlsx"

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val, 4)
    return str(val).strip()

def parse_excel():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found at {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    wb_form = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

    machines = [
        {"id": "ext_1", "name": "Ekstrüder Hat 1", "type": "extruder"},
        {"id": "ext_2", "name": "Ekstrüder Hat 2", "type": "extruder"},
        {"id": "ext_3", "name": "Ekstrüder Hat 3", "type": "extruder"},
        {"id": "ext_4", "name": "Ekstrüder Hat 4", "type": "extruder"},
        {"id": "ext_5", "name": "Ekstrüder Hat 5", "type": "extruder"},
        {"id": "ext_6", "name": "Ekstrüder Hat 6", "type": "extruder"},
        {"id": "ext_7", "name": "Ekstrüder Hat 7", "type": "extruder"},
        {"id": "ext_8", "name": "Ekstrüder Hat 8", "type": "extruder"},
        {"id": "ext_9", "name": "Ekstrüder Hat 9", "type": "extruder"},
        {"id": "ext_g1", "name": "Ekstrüder Genel 1", "type": "extruder"},
        {"id": "ext_g2", "name": "Ekstrüder Genel 2", "type": "extruder"},
        {"id": "lev_1", "name": "Levha Hattı 1", "type": "levha"},
        {"id": "lev_2", "name": "Levha Hattı 2", "type": "levha"},
    ]

    products = [
        {"id": "p_80_80", "name": "80X80 Pervaz", "category": "pervaz", "door_ratio": 5.0},
        {"id": "p_50_80", "name": "50x80 Pervaz", "category": "pervaz", "door_ratio": 5.0},
        {"id": "k_100", "name": "100 mm Kasa", "category": "kasa", "door_ratio": 2.5},
        {"id": "k_140", "name": "140 mm Kasa", "category": "kasa", "door_ratio": 2.5},
        {"id": "seren", "name": "Seren", "category": "seren", "door_ratio": 3.5},
        {"id": "kitkat", "name": "Kitkat", "category": "diger", "door_ratio": 0},
        {"id": "kitkat_emb", "name": "Kitkat (emboslu)", "category": "diger", "door_ratio": 0},
        {"id": "levha_dbudak", "name": "D. Budak", "category": "levha", "door_ratio": 2.0},
        {"id": "levha_teak", "name": "Teak", "category": "levha", "door_ratio": 2.0},
    ]

    daily_data = {}

    day_sheets = [f"{d:02d} Ağu" for d in range(1, 32)]

    for d_idx, sname in enumerate(day_sheets, 1):
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        
        # Read Day Shift
        gunduz_emp = clean_val(ws.cell(row=6, column=2).value) or 10
        gunduz_hours = clean_val(ws.cell(row=6, column=12).value) or 12

        gunduz_extruders = []
        for r in range(9, 29):
            hat = clean_val(ws.cell(row=r, column=1).value)
            product = clean_val(ws.cell(row=r, column=2).value)
            length = clean_val(ws.cell(row=r, column=3).value)
            speed = clean_val(ws.cell(row=r, column=4).value)
            prod_kg = clean_val(ws.cell(row=r, column=5).value) or 0
            fire_kg = clean_val(ws.cell(row=r, column=6).value) or 0
            qty = clean_val(ws.cell(row=r, column=7).value) or 0
            sets = clean_val(ws.cell(row=r, column=8).value) or 0

            if product or prod_kg or qty:
                gunduz_extruders.append({
                    "row": r,
                    "hat": str(hat) if hat else "",
                    "product": str(product) if product else "",
                    "length": length,
                    "speed": speed,
                    "prod_kg": float(prod_kg) if isinstance(prod_kg, (int, float)) else 0,
                    "fire_kg": float(fire_kg) if isinstance(fire_kg, (int, float)) else 0,
                    "qty": int(qty) if isinstance(qty, (int, float)) else 0,
                    "sets": float(sets) if isinstance(sets, (int, float)) else 0,
                })

        # Levha gündüz
        gunduz_levha = []
        for r in range(37, 47):
            hat = clean_val(ws.cell(row=r, column=1).value)
            color = clean_val(ws.cell(row=r, column=2).value)
            width = clean_val(ws.cell(row=r, column=3).value)
            length = clean_val(ws.cell(row=r, column=4).value)
            m2_one = clean_val(ws.cell(row=r, column=5).value)
            qty = clean_val(ws.cell(row=r, column=6).value) or 0
            total_m2 = clean_val(ws.cell(row=r, column=7).value) or 0
            width_fire_cm = clean_val(ws.cell(row=r, column=8).value) or 0
            dead_fire_m2 = clean_val(ws.cell(row=r, column=9).value) or 0
            dead_fire_kg = clean_val(ws.cell(row=r, column=10).value) or 0
            total_kg = clean_val(ws.cell(row=r, column=11).value) or 0
            sets = clean_val(ws.cell(row=r, column=12).value) or 0

            if color or qty or total_kg:
                gunduz_levha.append({
                    "row": r,
                    "hat": str(hat) if hat else "Levha 1",
                    "color": str(color) if color else "",
                    "width": width,
                    "length": length,
                    "m2_one": m2_one,
                    "qty": int(qty) if isinstance(qty, (int, float)) else 0,
                    "total_m2": float(total_m2) if isinstance(total_m2, (int, float)) else 0,
                    "width_fire_cm": float(width_fire_cm) if isinstance(width_fire_cm, (int, float)) else 0,
                    "dead_fire_m2": float(dead_fire_m2) if isinstance(dead_fire_m2, (int, float)) else 0,
                    "dead_fire_kg": float(dead_fire_kg) if isinstance(dead_fire_kg, (int, float)) else 0,
                    "total_kg": float(total_kg) if isinstance(total_kg, (int, float)) else 0,
                    "sets": float(sets) if isinstance(sets, (int, float)) else 0,
                })

        # Night shift (Gece)
        gece_emp = clean_val(ws.cell(row=64, column=2).value) or 10
        gece_hours = clean_val(ws.cell(row=64, column=12).value) or 12

        gece_extruders = []
        for r in range(67, 87):
            hat = clean_val(ws.cell(row=r, column=1).value)
            product = clean_val(ws.cell(row=r, column=2).value)
            length = clean_val(ws.cell(row=r, column=3).value)
            speed = clean_val(ws.cell(row=r, column=4).value)
            prod_kg = clean_val(ws.cell(row=r, column=5).value) or 0
            fire_kg = clean_val(ws.cell(row=r, column=6).value) or 0
            qty = clean_val(ws.cell(row=r, column=7).value) or 0
            sets = clean_val(ws.cell(row=r, column=8).value) or 0

            if product or prod_kg or qty:
                gece_extruders.append({
                    "row": r,
                    "hat": str(hat) if hat else "",
                    "product": str(product) if product else "",
                    "length": length,
                    "speed": speed,
                    "prod_kg": float(prod_kg) if isinstance(prod_kg, (int, float)) else 0,
                    "fire_kg": float(fire_kg) if isinstance(fire_kg, (int, float)) else 0,
                    "qty": int(qty) if isinstance(qty, (int, float)) else 0,
                    "sets": float(sets) if isinstance(sets, (int, float)) else 0,
                })

        gece_levha = []
        for r in range(95, 105):
            hat = clean_val(ws.cell(row=r, column=1).value)
            color = clean_val(ws.cell(row=r, column=2).value)
            width = clean_val(ws.cell(row=r, column=3).value)
            length = clean_val(ws.cell(row=r, column=4).value)
            m2_one = clean_val(ws.cell(row=r, column=5).value)
            qty = clean_val(ws.cell(row=r, column=6).value) or 0
            total_m2 = clean_val(ws.cell(row=r, column=7).value) or 0
            width_fire_cm = clean_val(ws.cell(row=r, column=8).value) or 0
            dead_fire_m2 = clean_val(ws.cell(row=r, column=9).value) or 0
            dead_fire_kg = clean_val(ws.cell(row=r, column=10).value) or 0
            total_kg = clean_val(ws.cell(row=r, column=11).value) or 0
            sets = clean_val(ws.cell(row=r, column=12).value) or 0

            if color or qty or total_kg:
                gece_levha.append({
                    "row": r,
                    "hat": str(hat) if hat else "Levha 1",
                    "color": str(color) if color else "",
                    "width": width,
                    "length": length,
                    "m2_one": m2_one,
                    "qty": int(qty) if isinstance(qty, (int, float)) else 0,
                    "total_m2": float(total_m2) if isinstance(total_m2, (int, float)) else 0,
                    "width_fire_cm": float(width_fire_cm) if isinstance(width_fire_cm, (int, float)) else 0,
                    "dead_fire_kg": float(dead_fire_kg) if isinstance(dead_fire_kg, (int, float)) else 0,
                    "total_kg": float(total_kg) if isinstance(total_kg, (int, float)) else 0,
                    "sets": float(sets) if isinstance(sets, (int, float)) else 0,
                })

        # Scrap & Downtimes
        downtimes = []
        for r in range(51, 62):
            hat = clean_val(ws.cell(row=r, column=1).value)
            fire_reason = clean_val(ws.cell(row=r, column=2).value)
            fire_kg = clean_val(ws.cell(row=r, column=5).value) or 0
            down_reason = clean_val(ws.cell(row=r, column=6).value)
            down_min = clean_val(ws.cell(row=r, column=9).value) or 0
            desc = clean_val(ws.cell(row=r, column=10).value) or ""

            if fire_reason or down_reason or fire_kg or down_min:
                downtimes.append({
                    "shift": "Gündüz",
                    "hat": str(hat) if hat else "",
                    "fire_reason": str(fire_reason) if fire_reason else "",
                    "fire_kg": float(fire_kg) if isinstance(fire_kg, (int, float)) else 0,
                    "down_reason": str(down_reason) if down_reason else "",
                    "down_min": float(down_min) if isinstance(down_min, (int, float)) else 0,
                    "desc": str(desc) if desc else "",
                })
        
        for r in range(109, 120):
            hat = clean_val(ws.cell(row=r, column=1).value)
            fire_reason = clean_val(ws.cell(row=r, column=2).value)
            fire_kg = clean_val(ws.cell(row=r, column=5).value) or 0
            down_reason = clean_val(ws.cell(row=r, column=6).value)
            down_min = clean_val(ws.cell(row=r, column=9).value) or 0
            desc = clean_val(ws.cell(row=r, column=10).value) or ""

            if fire_reason or down_reason or fire_kg or down_min:
                downtimes.append({
                    "shift": "Gece",
                    "hat": str(hat) if hat else "",
                    "fire_reason": str(fire_reason) if fire_reason else "",
                    "fire_kg": float(fire_kg) if isinstance(fire_kg, (int, float)) else 0,
                    "down_reason": str(down_reason) if down_reason else "",
                    "down_min": float(down_min) if isinstance(down_min, (int, float)) else 0,
                    "desc": str(desc) if desc else "",
                })

        daily_data[d_idx] = {
            "day": d_idx,
            "date": f"{d_idx:02d}.08.2026",
            "gunduz": {
                "employees": int(gunduz_emp) if isinstance(gunduz_emp, (int, float)) else 10,
                "hours": float(gunduz_hours) if isinstance(gunduz_hours, (int, float)) else 12,
                "extruders": gunduz_extruders,
                "levha": gunduz_levha,
            },
            "gece": {
                "employees": int(gece_emp) if isinstance(gece_emp, (int, float)) else 10,
                "hours": float(gece_hours) if isinstance(gece_hours, (int, float)) else 12,
                "extruders": gece_extruders,
                "levha": gece_levha,
            },
            "downtimes": downtimes
        }

    full_db = {
        "machines": machines,
        "products": products,
        "daily_data": daily_data
    }

    out_path = r"C:\Users\llboz\.gemini\antigravity\scratch\data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(full_db, f, ensure_ascii=False, indent=2)

    print(f"Successfully extracted {len(daily_data)} days into {out_path}")
    return full_db

if __name__ == "__main__":
    parse_excel()
